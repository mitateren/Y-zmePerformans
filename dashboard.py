# -*- coding: utf-8 -*-
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px
import plotly.io as pio
import numpy as np
import os
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import io

# Excel dosyasını oku
excel_file = 'yuzme_sonuclari.xlsx'
df = pd.read_excel(excel_file)

# Zaman sütununu datetime'a çevir
if 'Tarih' in df.columns:
    df['Tarih'] = pd.to_datetime(df['Tarih'], dayfirst=True, errors='coerce')

# Süreleri saniyeye çeviren fonksiyon
import re
def sure_to_seconds(s):
    if isinstance(s, str):
        if s in ['DNS', 'DSQ']:
            return np.nan
        match = re.match(r"(?:(\d+):)?(\d+)\.(\d+)", s.replace(':', '.'))
        if match:
            groups = match.groups('0')
            dakika = int(groups[0]) if groups[0] else 0
            saniye = int(groups[1])
            salise = int(groups[2])
            return dakika*60 + saniye + salise/100
    return np.nan
if 'Süre' in df.columns:
    df['Süre_saniye'] = df['Süre'].apply(sure_to_seconds)

# 1. Branş Bazlı Gelişim (Her sporcu-her branş)
plt.figure(figsize=(12,7))
for (sporcu, branş), grup in df.groupby(['Ad Soyad', 'Branş']):
    plt.plot(grup['Tarih'], grup['Süre_saniye'], marker='o', label=f"{sporcu} - {branş}")
plt.title('Sporcuların Branş Bazlı Gelişimi')
plt.xlabel('Tarih')
plt.ylabel('Süre (saniye)')
plt.legend(fontsize=7, bbox_to_anchor=(1.05, 1), loc='upper left')
plt.tight_layout()
plt.savefig('branş_gelisim.png')
plt.close()

# 2. Genel Gelişim (Her sporcu için tüm branşlar)
plt.figure(figsize=(12,7))
for sporcu, grup in df.groupby('Ad Soyad'):
    plt.plot(grup['Tarih'], grup['Süre_saniye'], marker='o', label=sporcu)
plt.title('Sporcuların Genel Gelişimi')
plt.xlabel('Tarih')
plt.ylabel('Süre (saniye)')
plt.legend(fontsize=8)
plt.tight_layout()
plt.savefig('genel_gelisim.png')
plt.close()

# 3. Cinsiyet ve Yaş Dağılımı
plt.figure(figsize=(6,4))
sns.countplot(x='Cinsiyet', data=df)
plt.title('Cinsiyet Dağılımı')
plt.savefig('cinsiyet_dagilimi.png')
plt.close()

plt.figure(figsize=(6,4))
sns.countplot(x='Doğum Yılı', data=df)
plt.title('Doğum Yılı Dağılımı')
plt.savefig('dogum_yili_dagilimi.png')
plt.close()

# 4. Branş Dağılımı
plt.figure(figsize=(10,5))
sns.countplot(y='Branş', data=df, order=df['Branş'].value_counts().index)
plt.title('Branşlara Göre Yarış Sayısı')
plt.savefig('brans_dagilimi.png')
plt.close()

# 5. En İyi Dereceler Tablosu (Her sporcu-branş için en iyi süre)
en_iyiler = df.dropna(subset=['Süre_saniye']).groupby(['Ad Soyad', 'Branş'])['Süre_saniye'].min().reset_index()
en_iyiler = en_iyiler.merge(df, on=['Ad Soyad', 'Branş', 'Süre_saniye'], how='left')
en_iyiler = en_iyiler[['Ad Soyad', 'Branş', 'Tarih', 'Süre', 'Süre_saniye']]
en_iyiler.to_excel('en_iyi_dereceler.xlsx', index=False)

# 6. Diskalifiye (DSQ) ve DNS Analizi
plt.figure(figsize=(6,4))
dsqs = df[df['Süre']=='DSQ'].groupby('Ad Soyad').size()
dnss = df[df['Süre']=='DNS'].groupby('Ad Soyad').size()
plt.bar(dsqs.index, dsqs.values, label='DSQ')
plt.bar(dnss.index, dnss.values, bottom=dsqs.reindex(dnss.index, fill_value=0).values, label='DNS')
plt.title('Sporcu Bazında DSQ ve DNS Sayıları')
plt.ylabel('Yarış Sayısı')
plt.xticks(rotation=90)
plt.legend()
plt.tight_layout()
plt.savefig('dsq_dns.png')
plt.close()

# 7. Plotly ile interaktif dashboard (HTML)
fig = px.line(df.dropna(subset=['Süre_saniye']), x='Tarih', y='Süre_saniye', color='Ad Soyad', line_dash='Branş',
              title='Tüm Sporcuların Branş ve Zamana Göre Gelişimi')
pio.write_html(fig, file='interaktif_dashboard.html', auto_open=False)

print("Dashboard ve tüm grafikler başarıyla oluşturuldu!\n- branş_gelisim.png\n- genel_gelisim.png\n- cinsiyet_dagilimi.png\n- dogum_yili_dagilimi.png\n- brans_dagilimi.png\n- dsq_dns.png\n- en_iyi_dereceler.xlsx\n- interaktif_dashboard.html")

# Dashboard özet ve grafiklerini Excel'e ekle
excel_dashboard = 'dashboard_rapor.xlsx'
wb = Workbook()
ws = wb.active
ws.title = 'Özet'

# Genel özet
ws['A1'] = 'Toplam Sporcu'
ws['B1'] = df['Ad Soyad'].nunique()
ws['A2'] = 'Toplam Yarış'
ws['B2'] = len(df)
ws['A3'] = 'Farklı Branş Sayısı'
ws['B3'] = df['Branş'].nunique()
ws['A4'] = 'Farklı Kulüp Sayısı'
ws['B4'] = df['Yarış Adı'].nunique() if 'Yarış Adı' in df.columns else ''

# Cinsiyet dağılımı
ws['A6'] = 'Cinsiyet Dağılımı'
for i, (cins, say) in enumerate(df['Cinsiyet'].value_counts().items(), start=7):
    ws[f'A{i}'] = cins
    ws[f'B{i}'] = say

# Branş dağılımı
ws2 = wb.create_sheet('Branş Dağılımı')
for i, (brans, say) in enumerate(df['Branş'].value_counts().items(), start=1):
    ws2[f'A{i}'] = brans
    ws2[f'B{i}'] = say

# En iyi dereceler
en_iyiler.to_excel('en_iyi_dereceler.xlsx', index=False)
ws3 = wb.create_sheet('En İyi Dereceler')
for r in en_iyiler.itertuples(index=False):
    ws3.append(r)

# Grafik ekleme fonksiyonu
def add_image_to_sheet(ws, img_path, cell):
    if os.path.exists(img_path):
        img = XLImage(img_path)
        ws.add_image(img, cell)

# Grafikler için ayrı sayfalar
ws4 = wb.create_sheet('Branş Gelişim Grafiği')
add_image_to_sheet(ws4, 'branş_gelisim.png', 'A1')
ws5 = wb.create_sheet('Genel Gelişim Grafiği')
add_image_to_sheet(ws5, 'genel_gelisim.png', 'A1')
ws6 = wb.create_sheet('Cinsiyet Dağılımı')
add_image_to_sheet(ws6, 'cinsiyet_dagilimi.png', 'A1')
ws7 = wb.create_sheet('Branş Dağılımı Grafik')
add_image_to_sheet(ws7, 'brans_dagilimi.png', 'A1')
ws8 = wb.create_sheet('DSQ-DNS Analizi')
add_image_to_sheet(ws8, 'dsq_dns.png', 'A1')

wb.save(excel_dashboard)
print('dashboard_rapor.xlsx dosyasına özet ve grafikler eklendi.')

# Verileri yuzme_sonuclari.json olarak kaydet
df.to_json('yuzme_sonuclari.json', orient='records', force_ascii=False, indent=2)
print('Veriler yuzme_sonuclari.json olarak kaydedildi.') 